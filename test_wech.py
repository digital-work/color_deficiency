'''
Created on 24. feb. 2014

@author: joschua
'''

from colordeficiency import *
#from tools import makeComparisonFig, makeSubplots, convertToLuminanceImage, mean_confidence_interval
import os
import sys
import subprocess
import colordeficiency.settings
import time
import colour
import random
import numpy
#from openpyxl import Workbook

from openpyxl import Workbook, load_workbook

    
def writeArrayToXLSX(dict,filename):
    
    wb = Workbook()   
    ws = wb.active
    i=0
    for key in dict:
        ws.title = key
        array_tmp = dict[key]  
        for r in array_tmp:
            ws.append(list(r))
        if i<len(dict)-1:
            ws = wb.create_sheet()
        i += 1
    
    wb.save(filename = filename)


def getStatsFromFilename(filename):
    """
    Returns the information of the file according to its filename:
    XXX                0|1            YY                                0|1           ZZ                             0|1|2|3
    ID of the image    Daltonized?    ID of the daltonization method    Simulated?    ID of the simulation method    Color deficiency type
    """
    
    dict = {}
    
    if (len(filename)!=10) or (filename.isdigit()!=True):
        print("Error: The filename is incorrect. It has to be 10-digits number.")
        return dict
    else:
        dict["img_id"] = int(filename[0:3])
        dict["dalt"] = bool(int(filename[3]))
        dict["dalt_id"] = int(filename[4:6])
        dict["sim"] = bool(int(filename[6]))
        dict["sim_id"] = int(filename[7:9])
        dict["coldef_type"] = int(filename[9])
        return dict
        
def setStatsToFilename(img_id,dalt,dalt_id,sim,sim_id,coldef_type):
    """
    Creates filename string with the according infomration status.
    XXX                0|1            YY                                0|1           ZZ                             0|1|2|3
    ID of the image    Daltonized?    ID of the daltonization method    Simulated?    ID of the simulation method    Color deficiency type
    """
    
    filename = ""
    ext = ".png"
    filename = str(img_id).zfill(3)+str(int(dalt))+str(dalt_id).zfill(2)+str(int(sim))+str(sim_id).zfill(2)+str(coldef_type)
    
    return filename+ext

def daltType2Int(daltonization_type):
    if daltonization_type=="anagnostopoulos":
        return 1
    elif daltonization_type == "kotera":
        return 2
    elif daltonization_type == "kuhn":
        return 3 
    elif daltonization_type == "huan":
        return 4
    elif daltonization_type == "dummy":
        return 99
    else:
        return 0

def simType2Int(simulation_type):
    if simulation_type=="vienot":
        return 1
    elif simulation_type == "vienot-adjusted":
        return 2
    elif simulation_type == "kotera":
        return 3 
    elif simulation_type == "brettel":
        return 4
    elif simulation_type == "dummy":
        return 99
    else:
        return 0

def read_csv_file(filename, pad=-numpy.inf):
    """
    Read a CSV file and return pylab array.

    Parameters
    ----------
    filename : string
        Name of the CSV file to read
    pad : float
        Value to pad for missing values.
    
    Returns
    -------
    csv_array : ndarray
        The content of the file plus padding.
    """
    
    wb = load_workbook(filename = filename, use_iterators = True)
    ws = wb.get_active_sheet() # ws is now an IterableWorksheet
    array = numpy.array([])
    for row in ws.iter_rows(): # it brings a new method: iter_rows()
        row_tmp = []
        for cell in row:
            row_tmp.append(cell.internal_value)
        if not array.size:
            array = numpy.array(row_tmp)
        else:
            array = numpy.vstack((array,numpy.array(row_tmp)))
    #print array
    array = numpy.array([x for x in array if x.any(None)])
    a =  array[:,0].astype(int)    
    b =  array[:,1].astype(bool)
    
    array = numpy.vstack((a,b))
    return array.transpose()

def makeXLSXForExperiment(options):
    """
    This method computes the XLSX files that are needed as input for the psychophysical experiments with PsychoPy2
    Options contains the following options:
        * path_in:     Source folder containing all the the images that should be included in the experiment.
        * extension:   File format of the files that should be included in the experiment. 
        * exp_in:      ID of the experiment for which the XLSX file should be computed. By default the XLSX is computed for all experiments.
    """
    
    n_exps = 2 # Number of experiments
    
    if 'path_in' in options:
        path_in = options['path_in']
    else:
        print("Error: No Source folder with images specified. Please call the function with the necessary location as path_in.")  
        return
    try:
        os.stat(os.path.join(path_in,'images'))
    except:
        os.mkdir(os.path.join(path_in,'images')) 
    
    if 'extension' in options:
        ext = options['extension']
    else:
        print("Caution: No extension specified for the files in the source folder. Taking the default extension png")
        ext = '.png'
        
    if 'exp_id' in options:
        exp_id = options['exp_id']
    else:
        print("Caution: No experiment ID chosen. We will compute the XLSX for all experiments together.")
        exp_id = 0

    if exp_id == 1:
        print("Making XlSX files for color deficiency simulation verification experiment")
        
        filename_xlsx_a = "mainTrialsA.xlsx"
        path_xlsx_a = os.path.join(path_in,filename_xlsx_a)
        filename_xlsx_b = "mainTrialsB.xlsx"
        path_xlsx_b = os.path.join(path_in,filename_xlsx_b)
        filename_xlsx_test = "testTrials.xlsx"
        path_xlsx_test = os.path.join(path_in,filename_xlsx_test)
        
        array_xlsx_a = numpy.array(['targetFile','controlFile','sim_id','coldef_type','origFile'])
        array_xlsx_b = numpy.array(['targetFile','controlFile','sim_id','coldef_type','origFile'])
        array_xlsx_test = numpy.array(['targetFile','controlFile','sim_id','coldef_type','origFile'])
        
        # Retrieve all image files in the source folder for further computation
        files = []
        files += [each for each in os.listdir(os.path.join(path_in,'images')) if each.endswith(ext)]
        n_files = len(files)
        
        n_simulations = len(settings.simulation_types)
        
        print("Including "+str(n_files)+" entries in the XLSX file ...")
        i = 0 # Counter for all files together
        j = 0 # Counter for all each simulated version of the color deficiency files
        k = 0 # Counter for test trials
        for file in files:
            i += 1
            #print os.path.basename(file)
            # Get name and general information about the image file
            filename_in_tmp = os.path.splitext(file)[0]
            sys.stdout.write(str(i).zfill(3) + "/"+str(n_files).zfill(3)+" --- "+ filename_in_tmp )
            dict_stats = getStatsFromFilename(filename_in_tmp)
            
            # Only process simulated images.
            if bool(dict_stats['sim']):
                j += 1
                
                path_out_tmp = os.path.join('images',file)
                
                # Get path to original file
                filename_orig_tmp = setStatsToFilename(dict_stats['img_id'],
                                                   dict_stats['dalt'],
                                                   dict_stats['dalt_id'],
                                                   False,
                                                   0,
                                                   0)
                path_orig_tmp = os.path.join('images',filename_orig_tmp)
                if (j%2 == 0):
                    array_xlsx_a = numpy.vstack((array_xlsx_a,[path_out_tmp,path_orig_tmp,settings.id2Sim[dict_stats['sim_id']],settings.id2ColDef[dict_stats['coldef_type']],path_orig_tmp]))
                    array_xlsx_b = numpy.vstack((array_xlsx_b,[path_orig_tmp,path_out_tmp,settings.id2Sim[dict_stats['sim_id']],settings.id2ColDef[dict_stats['coldef_type']],path_orig_tmp]))
                else:
                    array_xlsx_a = numpy.vstack((array_xlsx_a,[path_orig_tmp,path_out_tmp,settings.id2Sim[dict_stats['sim_id']],settings.id2ColDef[dict_stats['coldef_type']],path_orig_tmp]))
                    array_xlsx_b = numpy.vstack((array_xlsx_b,[path_out_tmp,path_orig_tmp,settings.id2Sim[dict_stats['sim_id']],settings.id2ColDef[dict_stats['coldef_type']],path_orig_tmp]))        
            else:
                k += 1
                sys.stdout.write("'s original.")
                
                path_orig_tmp = os.path.join('images',file)
                
                rand_sim_tmp = random.randint(1, n_simulations-1)
                if rand_sim_tmp == 3:
                    rand_coldef_type_tmp = 1 # For kotera only p can be chosen
                else:
                    rand_coldef_type_tmp = random.randint(1,2)
                filename_sim_rand_tmp = setStatsToFilename(dict_stats['img_id'],
                                                   dict_stats['dalt'],
                                                   dict_stats['dalt_id'],
                                                   True,
                                                   rand_sim_tmp,
                                                   rand_coldef_type_tmp)
                path_sim_rand_tmp = os.path.join('images',filename_sim_rand_tmp)
                if (k%2 == 0):
                    array_xlsx_test = numpy.vstack((array_xlsx_test,[path_sim_rand_tmp,path_orig_tmp,settings.id2Sim[rand_sim_tmp],settings.id2ColDef[rand_coldef_type_tmp],path_orig_tmp]))
                else:
                    array_xlsx_test = numpy.vstack((array_xlsx_test,[path_orig_tmp,path_sim_rand_tmp,settings.id2Sim[rand_sim_tmp],settings.id2ColDef[rand_coldef_type_tmp],path_orig_tmp]))
                    
            sys.stdout.write(".. OK\n")
        
        n_comparisons = numpy.shape(array_xlsx_a)[0]
        rounds = 35
        n_pauses = n_comparisons/rounds
        #print n_pauses
        header_a = array_xlsx_a[0,:]
        
        body_a = array_xlsx_a[1:n_comparisons+1]
        numpy.random.shuffle(body_a)
        array_xlsx_a[1:n_comparisons+1] = body_a
        
        filename_xlsx_counter_a = "mainTrialsA.xlsx"
        path_xlsx_counter_a = os.path.join(path_in,filename_xlsx_counter_a)
        array_xlsx_counter_a = numpy.array(['roundFile'])
        
        for i in range(0,n_pauses+1):
            array_xlsx_a_tmp = header_a
            array_xlsx_a_tmp = numpy.vstack([array_xlsx_a_tmp,array_xlsx_a[i*rounds+1:(i+1)*rounds+1,:]])
            
            filename_xlsx_a_tmp = os.path.join('rounds',"mainTrialsA"+str(i)+".xlsx")
            path_xlsx_a_tmp = os.path.join(path_in,filename_xlsx_a_tmp)
            dict_xlsx_a_tmp = {'mainTrials.csv':array_xlsx_a_tmp}
            writeArrayToXLSX(dict_xlsx_a_tmp,path_xlsx_a_tmp)
        
            array_xlsx_counter_a = numpy.vstack([array_xlsx_counter_a,filename_xlsx_a_tmp])
        
        dict_xlsx_counter_a = {'mainTrials.csv':array_xlsx_counter_a}
        writeArrayToXLSX(dict_xlsx_counter_a,path_xlsx_counter_a)
        
        # Making round files for b
        header_b = array_xlsx_a[0,:]
        
        body_b = array_xlsx_b[1:n_comparisons+1,:]
        numpy.random.shuffle(body_b)
        array_xlsx_b[1:n_comparisons+1,:] = body_b
        
        filename_xlsx_counter_b = "mainTrialsB.xlsx"
        path_xlsx_counter_b = os.path.join(path_in,filename_xlsx_counter_b)
        array_xlsx_counter_b = numpy.array(['roundFile'])
        
        for i in range(0,n_pauses+1):
            array_xlsx_b_tmp = header_b
            array_xlsx_b_tmp = numpy.vstack([array_xlsx_b_tmp,array_xlsx_b[i*rounds+1:(i+1)*rounds+1,:]])
            
            filename_xlsx_b_tmp = os.path.join('rounds',"mainTrialsB"+str(i)+".xlsx")
            path_xlsx_b_tmp = os.path.join(path_in,filename_xlsx_b_tmp)
            dict_xlsx_b_tmp = {'mainTrials.csv':array_xlsx_b_tmp}
            writeArrayToXLSX(dict_xlsx_b_tmp,path_xlsx_b_tmp)
        
            array_xlsx_counter_b = numpy.vstack([array_xlsx_counter_b,filename_xlsx_b_tmp])
        
        dict_xlsx_counter_b = {'mainTrials.csv':array_xlsx_counter_b}
        writeArrayToXLSX(dict_xlsx_counter_b,path_xlsx_counter_b)
        
        # Shuffling test
        header_test = array_xlsx_test[0,:]
        n_tests = 100
        
        body_test = array_xlsx_test[1:n_comparisons+1]
        numpy.random.shuffle(body_test)
        array_xlsx_test = header_test
        array_xlsx_test = numpy.vstack([array_xlsx_test,body_test[1:n_tests+1]])
        
        #dict_xlsx_a = {'mainTrials.csv':array_xlsx_a}
        
        #dict_xlsx_b = {'mainTrials.csv':array_xlsx_b}
        dict_xlsx_test = {'mainTrials.csv':array_xlsx_test}
        #writeArrayToXLSX(dict_xlsx_a,path_xlsx_a)
        #writeArrayToXLSX(dict_xlsx_b,path_xlsx_b)
        writeArrayToXLSX(dict_xlsx_test,path_xlsx_test)
        
        print("... finishing the computation of the XLSX files for the images.")
    
    elif exp_id == 2:
        print("Making XLSX files for color deficiency daltonization visual search experiments")
        
        filename_main_xlsx_a = 'mainTrialsA.xlsx'
        path_main_xlsx_a = os.path.join(path_in,filename_main_xlsx_a) # Path where XLSX file should be store. Supposed to be absolute path.
        array_main_xlsx_a = numpy.array(['trialFile','msgFile','keyFile'])
        filename_main_xlsx_b = 'mainTrialsB.xlsx'
        path_main_xlsx_b = os.path.join(path_in,filename_main_xlsx_b) # Path where XLSX file should be store. Supposed to be absolute path.
        array_main_xlsx_b = numpy.array(['trialFile','msgFile','keyFile'])
        
        filename_main_xlsx_test_a = 'mainTrialsTestA.xlsx'
        path_main_xlsx_test_a = os.path.join(path_in,filename_main_xlsx_test_a) # Path where XLSX file should be store. Supposed to be absolute path.
        array_main_xlsx_test_a = numpy.array(['trialFile','msgFile','keyFile'])
        filename_main_xlsx_test_b = 'mainTrialsTestB.xlsx'
        path_main_xlsx_test_b = os.path.join(path_in,filename_main_xlsx_test_b) # Path where XLSX file should be store. Supposed to be absolute path.
        array_main_xlsx_test_b = numpy.array(['trialFile','msgFile','keyFile'])
        
        # We assume that the sets are in a sudirectory called images
        path_images = os.path.join(path_in,'images')
        
        files_total = getAllXXXinPath(path_in,ext,with_subdirs=1)
        n_files = len(files_total)
        
        # Get all subdirectories of the image path
        subdirs = [x[0] for x in os.walk(path_images)] 
        
        print("Computing XLSX file(s) for "+str(n_files)+" image(s) ...")
        i = 0 # Counter for the images
        
        first = 1
        for subdir in subdirs:
            if first:
                # Skip the parent directory
                first = 0
            else:
                
                subdir_rel = os.path.basename(subdir)
                try:
                    os.stat(os.path.join(path_in,'images',subdir_rel))
                except:
                    os.mkdir(os.path.join(path_in,'images',subdir_rel)) 
                
                files = getAllXXXinPath(subdir,ext)#images_tmp += [each for each in os.listdir(subdir) if each.endswith('.png')]
                rel_path_tmp = os.path.join('images',os.path.basename(subdir)) # The information in the XLSX files should be stored as relative path
                if files:
                    j=0
                    n_subdir_files = len(files)
                    rand_files = [random.randint(1,n_subdir_files),random.randint(1,n_subdir_files),random.randint(1,n_subdir_files)]
                    # Include information to trial file and question file
                    array_main_xlsx_a = numpy.vstack((array_main_xlsx_a,[os.path.join(rel_path_tmp,'setTrialsA.xlsx'),os.path.join(rel_path_tmp,'questionA.txt'),os.path.join(rel_path_tmp,'keysA.txt')]))
                    answersA_tmp = read_csv_file(os.path.join(subdir,'answersA.xlsx'))
                    array_main_xlsx_b = numpy.vstack((array_main_xlsx_b,[os.path.join(rel_path_tmp,'setTrialsB.xlsx'),os.path.join(rel_path_tmp,'questionB.txt'),os.path.join(rel_path_tmp,'keysB.txt')]))
                    #print (os.path.join(subdir,'answersB.xlsx'))
                    answersB_tmp = read_csv_file(os.path.join(subdir,'answersB.xlsx'))
                    
                    array_main_xlsx_test_a = numpy.vstack((array_main_xlsx_test_a,[os.path.join(rel_path_tmp,'setTrialsTestA.xlsx'),os.path.join(rel_path_tmp,'questionA.txt'),os.path.join(rel_path_tmp,'keysA.txt')]))
                    answersTestA_tmp = read_csv_file(os.path.join(subdir,'answersA.xlsx'))
                    array_main_xlsx_test_b = numpy.vstack((array_main_xlsx_test_b,[os.path.join(rel_path_tmp,'setTrialsTestB.xlsx'),os.path.join(rel_path_tmp,'questionB.txt'),os.path.join(rel_path_tmp,'keysB.txt')]))
                    answersTestB_tmp = read_csv_file(os.path.join(subdir,'answersB.xlsx'))
                    
                    filename_set_xlsx_a = 'setTrialsA.xlsx'
                    path_set_xlsx_a = os.path.join(subdir,filename_set_xlsx_a) # Path to directory where the XLSX file should be stored. Should be absolute
                    array_set_xlsx_a = numpy.array(['stimFile','corrAns','dalt_id','coldef_type'])
                    filename_set_xlsx_b = 'setTrialsB.xlsx'
                    path_set_xlsx_b = os.path.join(subdir,filename_set_xlsx_b) # Path to directory where the XLSX file should be stored. Should be absolute
                    array_set_xlsx_b = numpy.array(['stimFile','corrAns','dalt_id','coldef_type'])
                    
                    filename_set_xlsx_test_a = 'setTrialsTestA.xlsx'
                    path_set_xlsx_test_a = os.path.join(subdir,filename_set_xlsx_test_a) # Path to directory where the XLSX file should be stored. Should be absolute
                    array_set_xlsx_test_a = numpy.array(['stimFile','corrAns','dalt_id','coldef_type'])
                    filename_set_xlsx_test_b = 'setTrialsTestB.xlsx'
                    path_set_xlsx_test_b = os.path.join(subdir,filename_set_xlsx_test_b) # Path to directory where the XLSX file should be stored. Should be absolute
                    array_set_xlsx_test_b = numpy.array(['stimFile','corrAns','dalt_id','coldef_type'])
                    
                    #print answersA_tmp
                    #print files 
                    for file in files:
                        i += 1
                        j += 1
                        
                        filename_in_tmp = os.path.splitext(file)[0]
                        sys.stdout.write(str(i).zfill(3) + "/"+str(n_files).zfill(3)+" --- "+ filename_in_tmp )
                        dict_stats = getStatsFromFilename(filename_in_tmp)
                        
                        #print dict_stats['img_id']
                        
                        #print answersA_tmp[:,0]==int(dict_stats['img_id'])
                        # Get correct answer for current image ID
                        corr_answerA_tmp = bool(answersA_tmp[answersA_tmp[:,0]==int(dict_stats['img_id']),1][0])
                        corr_answerA_tmp = 'left' if corr_answerA_tmp else 'right'
                        corr_answerB_tmp = bool(answersB_tmp[answersB_tmp[:,0]==int(dict_stats['img_id']),1][0])
                        corr_answerB_tmp = 'left' if corr_answerB_tmp else 'right'

                        path_orig = os.path.join(rel_path_tmp,file) # Has to be relative
                        array_set_xlsx_a = numpy.vstack((array_set_xlsx_a,[path_orig,corr_answerA_tmp,settings.id2Dalt[dict_stats['dalt_id']],settings.id2ColDef[dict_stats['coldef_type']]]))
                        array_set_xlsx_b = numpy.vstack((array_set_xlsx_b,[path_orig,corr_answerB_tmp,settings.id2Dalt[dict_stats['dalt_id']],settings.id2ColDef[dict_stats['coldef_type']]]))
                        
                        if j in rand_files:
                            if dict_stats['dalt_id']==99:
                                sys.stdout.write('.notestdummy.')
                                rand_files.append(j+7)
                            else:
                                corr_answerTestA_tmp = bool(answersTestA_tmp[answersTestA_tmp[:,0]==int(dict_stats['img_id']),1][0])
                                corr_answerTestA_tmp = 'left' if corr_answerTestA_tmp else 'right'
                                array_set_xlsx_test_a = numpy.vstack((array_set_xlsx_test_a,[path_orig,corr_answerTestA_tmp,settings.id2Sim[dict_stats['dalt_id']],settings.id2ColDef[dict_stats['coldef_type']]]))
                                
                                corr_answerTestB_tmp = bool(answersTestB_tmp[answersTestB_tmp[:,0]==int(dict_stats['img_id']),1][0])
                                corr_answerTestB_tmp = 'left' if corr_answerTestB_tmp else 'right'
                                array_set_xlsx_test_b = numpy.vstack((array_set_xlsx_test_b,[path_orig,corr_answerTestB_tmp,settings.id2Sim[dict_stats['dalt_id']],settings.id2ColDef[dict_stats['coldef_type']]]))
                                
                        
                        sys.stdout.write(" . OK\n")
                        
                    dict_set_xlsx_a = {'setTrials.csv': array_set_xlsx_a}
                    writeArrayToXLSX(dict_set_xlsx_a, path_set_xlsx_a)
                    dict_set_xlsx_b = {'setTrials.csv': array_set_xlsx_b}
                    writeArrayToXLSX(dict_set_xlsx_b, path_set_xlsx_b)
                    
                    dict_set_xlsx_test_a = {'setTrials.csv': array_set_xlsx_test_a}
                    writeArrayToXLSX(dict_set_xlsx_test_a, path_set_xlsx_test_a)
                    dict_set_xlsx_test_b = {'setTrials.csv': array_set_xlsx_test_b}
                    writeArrayToXLSX(dict_set_xlsx_test_b, path_set_xlsx_test_b)
                        
        dict_main_xlsx_a = {'mainTrials.csv': array_main_xlsx_a}
        writeArrayToXLSX(dict_main_xlsx_a,path_main_xlsx_a)
        dict_main_xlsx_b = {'mainTrials.csv': array_main_xlsx_b}
        writeArrayToXLSX(dict_main_xlsx_b,path_main_xlsx_b)
        
        dict_main_xlsx_test_a = {'mainTrials.csv': array_main_xlsx_test_a}
        writeArrayToXLSX(dict_main_xlsx_test_a,path_main_xlsx_test_a)
        dict_main_xlsx_test_b = {'mainTrials.csv': array_main_xlsx_test_b}
        writeArrayToXLSX(dict_main_xlsx_test_b,path_main_xlsx_test_b)
        
        print("... finishing the computation of the XLSX file(s) for images.")
        
    else:
        print("Caution: We will compute the XLSX for all experiments together.")
        
        options_adj = options.copy()
        
        options['exp_id'] = 1
        makeXLSXForExperiment(options_adj)
        
        options['exp_id'] = 2
        makeXLSXForExperiment(options_adj)

def getAllXXXinPath(path,ext,with_subdirs=0):
    """
    Returns a list with all the files that have the specified extension in the path
    """
    if with_subdirs:
        files = []
        subdirs = [x[0] for x in os.walk(path)]
        for subdir in subdirs:
            files += [each for each in os.listdir(subdir) if each.endswith(ext)]
            
    else:
        files = []
        files += [each for each in os.listdir(path) if each.endswith(ext)]
    
    return files



def test16(options):
    """
    Preparing images for the color deficiency daltonization visual search experiments (ID: 2).
    Dict contains the following options:
        * path_in:                 Source folder containing all the images that are about to be simulated.
        * extension:               File format of hte files that should be simulated in the folder. PNG by default.
        * path_out:                Destination folder, in which all the simulated images should be saved. path_in location by default.
        * daltonization_types:     Daltonization methods that are used to daltonize the images.
        * colordeficiency_types:   Types of color deficiency for which the images should be simulated.
    """
    
    if 'path_in' in options:
        path_in = options['path_in'] # Should be specified as absolute path
    else:
        print("Error: No Source folder with images specified. Please call the function with the necessary location as path_in.")  
        return
    
    if 'extension' in options:
        ext = options['extension']
    else:
        print("Caution: No extension specified for the files in the source folder. Taking the default extension png")
        ext = '.png'
    
    if 'path_out' in options:
        path_out = options['path_out']
    else:
        print("Caution: No destination folder specified. Taking the same directory as the source folder: "+ path_in)
        path_out = path_in
    try:
        os.stat(os.path.join(path_out,'images'))
    except:
        os.mkdir(os.path.join(path_out,'images')) 
    
    if 'daltonization_types' in options:
        daltonization_types = options['daltonization_types']
    else:
        print("Caution: No daltonization types chosen. Choosing default daltonization types instead: "+str(settings.daltonization_types))
        daltonization_types = settings.daltonization_types
    
    if 'coldef_types' in options:
        coldef_types = options['coldef_types']
    else:
        print("Caution: No color deficiency types chose. Choosing default color deficiencies instead "+str(settings.coldef_types))
        coldef_types = settings.coldef_types
    
    # Get number of all files to be computed in the source folder and its subdirectories for the counter
    files_total = getAllXXXinPath(path_in,ext,with_subdirs=1)
    n_files = len(files_total)
    
    # Get all subdirectories of the source path
    subdirs = [x[0] for x in os.walk(os.path.join(path_in,'images'))] 
    
    print("Starting with the computation of the "+str(len(coldef_types))+"x"+str(len(daltonization_types))+" daltonization for the "+str(n_files)+" image(s) ...")
    i = 0 # Counter for counting the files in the source folder
    first = 1
    for subdir in subdirs:
        if first:
            # Skip the parent directory
            first = 0
        else:
            subdir_rel = os.path.basename(subdir)
            try:
                os.stat(os.path.join(path_out,'images',subdir_rel))
            except:
                os.mkdir(os.path.join(path_out,'images',subdir_rel)) 
            
            # Retrieve all image files in the subdirectory folder of the source folder
            files = getAllXXXinPath(os.path.join(path_in,subdir),ext)#images_tmp += [each for each in os.listdir(subdir) if each.endswith('.png')]
 
            if files: # Only compute folders, where there are image files
                
                for file in files:
                    i += 1
                    
                    # Get name and general infomration about the image file.
                    filename_in_tmp = os.path.splitext(file)[0]
                    sys.stdout.write(str(i).zfill(3) + "/"+str(n_files).zfill(3)+" --- "+ filename_in_tmp )
                    dict_stats = getStatsFromFilename(filename_in_tmp)
                    
                    # Only process original images. Don't process daltonized or simulated images
                    if not bool(dict_stats['dalt']):
                        
                        schon_dummy = 1
                        
                        size = 900,900
                        #print subdir
                        img_in = Image.open(os.path.join(subdir,file))
                        if numpy.shape(img_in)[0:2] != size:
                            img_in.thumbnail(size,Image.ANTIALIAS)
                        
                        #img_in.thumbnail((256,256))
                        
                        # Save original image in the destination folder
                        path_out_orig = os.path.join(os.path.join(path_out,'images',subdir_rel),file)
                        img_in.save(path_out_orig)
                        
                        #img_in = Image.open(path_out_orig)
                        
                        for daltonization_type in daltonization_types:
                            sys.stdout.write(" "+daltonization_type+".")
                            
                            for coldef_type in coldef_types:
                                
                                if (daltonization_type == 'dummy') and not schon_dummy:
                                    sys.stdout.write(coldef_type+".schunngmacht.")
                                    pass
                                else:
                                    
                                    if (daltonization_type == 'dummy') and schon_dummy:
                                        schon_dummy = 0
                                    
                                    sys.stdout.write(coldef_type+".")
                                    
                                    try:
                                        # Daltonize image and save it in the destination folder
                                        img_dalt = daltonize(img_in,{'coldef_type': coldef_type, 'daltonization_type': daltonization_type})
                                        filename_out_tmp = setStatsToFilename(dict_stats['img_id'],
                                                                  True,
                                                                  daltType2Int(daltonization_type),
                                                                  dict_stats['sim'],
                                                                  dict_stats['sim_id'],
                                                                  coldefType2Int(coldef_type))
                                        path_out_tmp = os.path.join(path_out,'images',subdir_rel,filename_out_tmp)
                                        img_dalt.save(path_out_tmp)
                                    except:
                                        print("Error: Could not simulate file: " + str(file) + " with " + str(daltonization_type))
                                        pass
                                
                    else:
                        
                        sys.stdout.write(" Image is already daltonized")
                        
                    sys.stdout.write(".. OK\n")
                    
    print("... finishing the computation of the daltonizations for the images.")

#path_in = '/Users/thomas/Dropbox/00_NZT/01_PhD/03_Experiments/00_Color-Deficiency-Evaluation-Experiments-Sessions/Session-1/01_experiments/visual-search/'
#path_out = "."      
#makeXLSXForExperiment(path_in, path_out,2,{})
    
    
    
def test15(options):
    """
    Preparing images for the color deficiency simulation verification experiment (ID: 1).
    Dict contains the following options:
        * path_in:                 Source folder containing all the images that are about to be simulated.
        * extension:               File format of hte files that should be simulated in the folder. PNG by default.
        * path_out:                Destination folder, in which all the simulated images should be saved. path_in location by default.
        * simulation_types:        Simulation methods that are used to simulated the images.
        * colordeficiency_types:   Types of color deficiency for which the images should be simulated.
    """
    
    
    if 'path_in' in options:
        path_in = options['path_in'] # Should be specified as absolute path
    else:
        print("Error: No Source folder with images specified. Please call the function with the necessary location as path_in.")  
        return
    
    if 'extension' in options:
        ext = options['extension']
    else:
        print("Caution: No extension specified for the files in the source folder. Taking the default extension png")
        ext = '.png'
    
    if 'path_out' in options:
        path_out = options['path_out']
    else:
        print("Caution: No destination folder specified. Taking the same directory as the source folder: "+ path_in)
        path_out = path_in
    try:
        os.stat(os.path.join(path_out,'images'))
    except:
        os.mkdir(os.path.join(path_out,'images')) 
    
    if 'simulation_types' in options:
        simulation_types = options['simulation_types']
    else:
        print("Caution: No simulation types chosen. Choosing default color deficiencies instead: "+str(settings.simulation_types))
        simulation_types = settings.simulation_types
    
    if 'coldef_types' in options:
        coldef_types = options['coldef_types']
    else:
        print("Caution: No color deficicion types chose. Choosing default color deficiencies instead " + str(settings.coldef_types))
        coldef_types = settings.coldef_types
    
    # Retrieve all image files in the source folder for further computation
    files = getAllXXXinPath(os.path.join(path_in,'images'),ext)#+= [each for each in os.listdir(path_in) if each.endswith(ext)]
    n_files = len(files)
    
    print("Starting with the computation of the "+str(len(coldef_types))+"x"+str(len(simulation_types))+" simulations for the "+str(n_files)+" image(s) ...")
    i = 0 # Counter for counting the files in the source folder
    for file in files:
        i += 1
        
        # Get name and general information about the image file.
        filename_in_tmp = os.path.splitext(file)[0]
        sys.stdout.write(str(i).zfill(3) + "/"+str(n_files).zfill(3)+" --- "+ filename_in_tmp )
        dict_stats = getStatsFromFilename(filename_in_tmp)
        
        # Only process original images. Don't process daltonized or simulated images.
        if not bool(dict_stats['sim']):
            
            schon_dummy = 1
            schon_kotera = 1
            size = 900,900
            img_in = Image.open(os.path.join(path_in,'images',file))
            #img_size = numpy.shape(img_in)[0:2]
            #print img_size
            if numpy.shape(img_in)[0:2] != size:
                img_in.thumbnail(size,Image.ANTIALIAS)
            #img_in.thumbnail((256,256))
            
            # Save original image in the destination folder
            path_out_orig = os.path.join(path_out,'images',file)
            img_in.save(path_out_orig)
            
            #img_in = Image.open(path_out_orig)
            
            for simulation_type in simulation_types:
                sys.stdout.write(" "+simulation_type+".")
                
                for coldef_type in coldef_types:
                    
                    if (simulation_type == 'dummy') and not schon_dummy:
                        sys.stdout.write(coldef_type+".schunngmacht.")
                    elif (simulation_type == 'kotera') and not schon_kotera:
                        sys.stdout.write(coldef_type+".schunngmacht.")
                    else:
                        
                        if (simulation_type == 'dummy') and schon_dummy:
                            schon_dummy = 0
                            
                        if (simulation_type == 'kotera') and schon_kotera:
                            schon_kotera = 0
                        
                        sys.stdout.write(coldef_type+".")
                        
                        # Simulate image and save it in the destination folder
                        try:
                            img_sim = simulate(simulation_type,img_in,coldef_type)
                            filename_out_tmp = setStatsToFilename(dict_stats['img_id'],
                                                                  dict_stats['dalt'],
                                                                  dict_stats['dalt_id'],
                                                                  True,
                                                                  simType2Int(simulation_type),
                                                                  coldefType2Int(coldef_type)) # Creates filenmae such that information about the image is decoded in the filename
                            path_out_tmp = os.path.join(path_out,'images',filename_out_tmp)
                            img_sim.save(path_out_tmp)
                        except:
                            print("Error: Could not simulate file: " + str(file) + " with " + str(simulation_type))
                            pass
                    
        else:
            
            sys.stdout.write("'s already simulated")
          
        sys.stdout.write(".. OK\n")
   
    print("... finishing the computation of the simulations for the images.")

def test17():
    
    print("Starting simulations")
    options = {'path_in': os.path.join('/Users/thomas/Desktop/test/source/','exp1'),
               'path_out': '/Users/thomas/Dropbox/00_NZT/01_PhD/03_Experiments/00_Color-Deficiency-Evaluation-Experiments-Sessions/00_Session/01_experiments/sample-2-match/',
               'coldef_types': ['p','d']}
   
    #            'path_out': '/Users/thomas/Dropbox/00_NZT/01_PhD/03_Experiments/00_Color-Deficiency-Evaluation-Experiments-Sessions/Session-1/01_experiments/sample-2-match/'}
    
    #test15(options)
    
    
    options = {'path_in': '/Users/thomas/Dropbox/00_NZT/01_PhD/03_Experiments/00_Color-Deficiency-Evaluation-Experiments-Sessions/00_Session/01_experiments/sample-2-match/',
               'exp_id':1}
    try:
        makeXLSXForExperiment(options)
        pass
    except:
        print("Error: Problems with the making of the XLSX file for experiment 1.")
    
    options = {'path_in': os.path.join('/Users/thomas/Desktop/test/source/','exp2'),
               'path_out': '/Users/thomas/Dropbox/00_NZT/01_PhD/03_Experiments/00_Color-Deficiency-Evaluation-Experiments-Sessions/00_Session/01_experiments/visual-search/',
               'coldef_types': ['p','d'],
               'daltonization_types': ['anagnostopoulos','kuhn','dummy','kotera'],
               'max_num_quant': 128}
    
    #test16(options)
    
    options = {'path_in': '/Users/thomas/Dropbox/00_NZT/01_PhD/03_Experiments/00_Color-Deficiency-Evaluation-Experiments-Sessions/00_Session/01_experiments/visual-search/',
               'exp_id':2}
    try:
        makeXLSXForExperiment(options)
        pass
    except:
        print("Error: Problems with the making of the XLSX file for experiment 2.")
    
def test18():
    path = '/Users/thomas/Dropbox/00_NZT/01_PhD/03_Experiments/00_Color-Deficiency-Evaluation-Experiments-Sessions/00_Session/01_experiments/visual-search/'
    #files =  getAllXXXinPath(path,'.png',with_subdirs=1)
    files = []
    subdirs = [x[0] for x in os.walk(os.path.join(path,'images'))] 
    for subdir in subdirs:
        #print subdir
        files_tmp = getAllXXXinPath(subdir,'.png')
        files_tmp = [os.path.join(subdir,file) for file in files_tmp]
        files = numpy.hstack([files,files_tmp])
    print(files)
    
    size = 900,900
        
    for file in files:
        img_tmp = Image.open(file)
        #print subdir
        img_in = Image.open(os.path.join(subdir,file))
        if numpy.shape(img_in)[0:2] != size:
            img_in.thumbnail(size,Image.ANTIALIAS)
        img_tmp.save(file)
        
    print('Done')
    
#test17()

def test14():
    
    size = 256,256
    
    path = "images/database/02_eksperiment_vis-sear/00_top/"
    fileList = os.listdir(path)
    
    for file in fileList:
        if "thumb" not in file:
            try:
                img_in = Image.open(os.path.join(path,file))
                img_in.thumbnail(size, Image.ANTIALIAS)
                
                name =  os.path.splitext(file)[0]
                img_in.save(os.path.join(path,name+"-thumb.png"))
            except IOError:
                pass
                    

def test13():
    img_in = Image.open('images/example31.jpg')

    
    
    t = time.time()
    img_out_1 = convertToLuminanceImage(img_in)
    #img_out_1.show()    
    elapsed = time.time() - t
    print(elapsed)
     
    t = time.time()
    img_out_2 = img_in.copy()
    img_out_2 = img_out_2.convert('LA')#convertToLuminanceImage(img_in)
    #img_out_2.show() 
    elapsed = time.time() - t
    print(elapsed)
    
    diff_arr = colour.metric.dE_00(colour.data.Data(colour.space.srgb, img_out_1.convert('RGB')), colour.data.Data(colour.space.srgb, img_out_2.convert('RGB')))
    img_dif = Image.fromarray(diff_arr)
    
    m,n = numpy.shape(diff_arr)
    diff_vect = numpy.reshape(diff_arr, m*n)
    print(mean_confidence_interval(diff_vect))
    img_dif.show()
    import matplotlib.pyplot as pyplot
    pyplot.boxplot( diff_vect)
    pyplot.show()

def test11():
    
    #os.system("./stress -i ./images/example13.png")
    os.environ['PATH'] = os.environ['PATH'] + ":/usr/local/bin"
    print(os.environ['PATH'])
    os.system("./stress -i ./images/example13.jpg -o ./test13.png -g -ns 1 -ni 200")
    #subprocess.call("./stress -i ./images/example11.jpg -o ./test1.png -g -ns 1 -ni 200")

def test12():
    
    img_in = Image.open("images/example7.png")
    img_in.show()
    img_sim_p = simulate("kotera",img_in,"p")
    img_sim_p.show()
    
    img_dif = Image.fromarray(colour.metric.dE_00(colour.data.Data(colour.space.srgb, img_in), colour.data.Data(colour.space.srgb, img_sim_p)))
    img_dif.show()
    

def test10():
    """
    Testing yoshi_c2g algorithm
    """
    coldef_types = settings.coldef_types#[p,d,t]
    simulation_type = settings.default['coldef_type']
    daltonization_types = ['yoshi_c2g','yoshi_c2g_only']
    size = 1024,1024
    
    enhances = [1,0]
    pts = 5
    its = 100
    
    names = ["images/database/trit9.jpg","images/database/ber3.jpg","images/database/wrest26.jpg","images/database/nat2.jpg","images/database/pap11.jpg"]
    #name = "images/example1.jpg"
    #name = "images/database/wrest30.jpg"
    for name in names:
        print(name)
        img_in = Image.open(name)
        for daltonization_type in daltonization_types:
            print(daltonization_type)
            for enhance in enhances:
                for coldef_type in coldef_types:
                    img_in.thumbnail(size)
                    img_in_sim = simulate(simulation_type,img_in,coldef_type)
                    
                    dict_dalt = {'daltonization_type':daltonization_type, 'coldef_type':coldef_type}
                    dict_dalt.update({'enhance':enhance, 'pts':pts, 'its':its})
                    img_out = daltonize(img_in,dict_dalt)
                    img_out_sim = simulate(simulation_type,img_out,coldef_type)
                    
                    size_plts = (2,2)
                    imgs = [{'img_in':img_in,'title':'Orig. image','fontsize':40.}]
                    imgs.append({'img_in':img_in_sim,'title':'Orig. sim:\"'+simulation_type+"\" | "+coldef_type,'fontsize':40.})
                    
                    if enhance:
                        imgs.append({'img_in':img_out,'title':'dalt:\"'+daltonization_type+"\" +\n pts:"+str(pts)+" its:"+str(its)+" | "+coldef_type,'fontsize':40.})
                        imgs.append({'img_in':img_out_sim,'title':'sim:\"'+simulation_type+"\' | dalt:\'"+daltonization_type+"\' +\n pts:"+str(pts)+" its:"+str(its)+" | "+coldef_type,'fontsize':40.})
                    else:
                        imgs.append({'img_in':img_out,'title':'dalt:\"'+daltonization_type+"\" | "+coldef_type,'fontsize':40.})
                        imgs.append({'img_in':img_out_sim,'title':'sim:\"'+simulation_type+"\' | dalt:\'"+daltonization_type+"\' | "+coldef_type,'fontsize':40.})
                    
                    if not enhance and daltonization_type=="yoshi_c2g":
                        output_name = './images/yoshi_c2g/subplots-'+os.path.basename(name)+"-"+coldef_type+"-"+daltonization_type+"-not-enhanced.png"
                    else:
                        output_name = './images/yoshi_c2g/subplots-'+os.path.basename(name)+"-"+coldef_type+"-"+daltonization_type+"-enhanced-pts-"+str(pts)+"-its-"+str(its)+".png"

                    options = {'output_name':output_name,'size_inches':settings.A1}
                    fig = makeSubplots(size_plts,imgs,options)
    
def test9():
    """
    pupsi dupsi
    """
    name = "example"
    best = [7] #33
    coldef_types = settings.coldef_types
    simulation_types = settings.simulation_types#[vienot,vienot_adjusted,kotera,brettel]
    size = 1024,1024
    
    for b in best:
        name_tmp = name+str(b)
        sys.stdout.write("Computing \'"+name_tmp+"\'.")       
        img_in = Image.open("images/buginbrettel/"+name_tmp+".png")
        img_in.thumbnail(size)     
        for simulation_type in simulation_types:
            sys.stdout.write(str(simulation_type).upper()+".")
            for coldef_type in coldef_types:
                sys.stdout.write(coldef_type+".")
                pylab.figure()
                pylab.subplot(121)
                pylab.title("Original image")
                pylab.axis("off")
                pylab.imshow(img_in)
                
                img_out = simulate(simulation_type,img_in,coldef_type)
                name_sim = "images/buginbrettel/tritanoper/"+name_tmp+"-"+simulation_type+"-"+coldef_type+"-sim.jpg"
                img_out.save(name_sim)
                
                pylab.subplot(122)
                pylab.title("Sim.: "+str(simulation_type)+" | ColDef.: "+str(coldef_type))
                pylab.imshow(img_out)
                pylab.axis("off")
                name_comparison = "images/buginbrettel/tritanoper/"+name_tmp+"-"+simulation_type+"-"+coldef_type+"-comparison.jpg"
                pylab.savefig(name_comparison)
                pylab.close()
                
                
        sys.stdout.write("ok.\n")
        #makeComparisonFig(img_in,coldef_types,simulation_types,name_tmp)   
        #pylab.savefig(name_tmp + "-comparison2.png")
    #pylab.show()
    sys.stdout.write("Done, gurl ;)")

#print test9.__doc__

def test8(options):
    """
    Test8: Compute color deficiencies of one or multiple input images plottet as subplots in one image.
    Input: -s: size - Resolution of the simulations in the final plot. 1024x1024 by default.
           -i: inputfiles - Name of the inputfiles
           -p: path - Path where the inputfils are located and where the figure will be saved
           -c: coldef_types - Types of color deficiencies being simulated. Protanopia by default.
           -u: simulation_types - Types of simulations used for simulation. Brettel by default.
    """
    if 'size' in options:
        s = int(options['size'])
    
        if s == 0:
            size = 1024,1024
        elif s == 1:
            size = 512,512
        elif s == 2:        
            size = 256,256#1024,1024#512,512
    else:
        print("Caution: No size chosen. Chose default 1024x1024 instead.")
        size == 1024,1024
    
    if 'inputfiles' in options:
        inputfiles = options['inputfiles']
        fileList = inputfiles.split(',')
    else:
        print("Error: No inputfile(s) chosen. Function can not be called.")
    
    print("Started [Size:%sx%s]" % size)
    
    if 'coldef_types' in options:
        coldef_types = options['coldef_types']
    else:
        coldef_types = ['p']
    
    if 'simulation_types' in options:
        simulation_types = options['simulation_types']
    else:
        simulation_types = ['brettel']        
    
    if 'path' in options:
        folder = options['path']
    else:
        folder = "images/database/02_eksperiment_vis-sear/Natur/00_top/ber3/"
    
    n_files = numpy.shape(fileList)
    i=0
    
    for name in fileList:
        #print name
        i = i+1
        #print name
        save_name = folder + name + "-comparison.png"
        if "comparison" not in name:# and not os.path.isfile(save_name):
            try: 
                img_in = Image.open(folder+name)
                img_in.thumbnail(size)
                
                sys.stdout.write(str(i) + "/"+str(n_files[0])+" --- ")
                
                makeComparisonFig(img_in, coldef_types,simulation_types,name)
                pylab.savefig(save_name)
                pylab.close()
                #img_in.show()
            except IOError:
                sys.stdout.write(str(i) + "/"+str(n_files[0])+" --- Error: Could not load image: \'" + name + "\'\n")
                
                pass
    print("Finished!")
    pylab.show()

def test7():
    
    name = "example"
    best = [10,15,17,18,19,33] #33
    coldef_types = ['t']
    simulation_types = ['brettel']
    size = 1024,1024
    
    for b in best:
        name_tmp = name+str(b)       
        img_in = Image.open("images/"+name_tmp+".jpg")
        img_in.thumbnail(size)        
        #img_in.show()
        
        makeComparisonFig(img_in,coldef_types,simulation_types,name_tmp)   
        pylab.savefig(name_tmp + "-comparison2.png")
    #pylab.show()
    
def test6():
    #best = [10,12,14,20,25,5,6] 
    #best = [101,102]
    best = [5]
    
    name = "example"
    coldef_types = ["p","d"]
    simulation_types = ["vienot","kotera"]
    daltonization_type = "anagnatopoulos"
    size = 256, 256
    
    #import pylab
    for b in best:
        print(1)
        for simulation_type in simulation_types:
            print(2)
            for coldef_type in coldef_types:
                print("starting with " + coldef_type)
                #pylab.figure()
                name_tmp = name+str(b)
                
                img_in = Image.open("images/egne/"+name_tmp+".jpg")
                
                img_in.thumbnail(size, Image.ANTIALIAS)
                
                img_in_sim = simulate(simulation_type,img_in,coldef_type)
                #                 img_in_sim.save("images/"+name_tmp+"_"+str(coldef_type)+"_sim-"+str(simulation_type)+".jpg")
                img_in_sim.show()
                #                 
                #                 img_dalt = daltonize(daltonization_type,img_in,coldef_type)
                #                 img_dalt.save("images/"+name_tmp+"_"+str(coldef_type)+"_dalt-"+str(daltonization_type)+".jpg")
                #                 img_dalt.show()
                #                 
                #                 img_dalt_sim = simulate(simulation_type,img_dalt,coldef_type)
                #                 img_dalt_sim.save("images/"+name_tmp+"_"+str(coldef_type)+"_dalt-"+str(daltonization_type)+"_sim-"+str(simulation_type)+".jpg")
                #                 img_dalt_sim.show()
                
                #pylab.savefig(name_tmp+"-"+coldef_type+"-costs-fig.png")
                #img_sim.show()
    pylab.show()

def test5():
    """
    Make example images to illustrate evaluation methods
    """
    
    best = [1,6,10,70] # Img 5 illustrates daltonization, img 10 illustrates the color deficiency verification experiments and , img 1 illustrates the daltonization evaluation experiment - visual search, img 60 illustrates the daltonization evaluation experiment - object recognition 
    
    name = "example"
    coldef_type = "d"
    simulation_type = "vienot-adjusted"
    daltonization_type = "kotera"
    size = 128, 128
    
    for b in best:
        name_tmp = name+str(b)
        
        img_in = Image.open("images/"+name_tmp+".jpg")
        
        #img_in.thumbnail(size, Image.ANTIALIAS)
        img_in.save("images/presentation/"+name_tmp+"orig.jpg", "JPEG")
        #img_in.show()
        img_in_sim = simulate(simulation_type, img_in, coldef_type)
        #img_in_sim.show()
        img_in_sim.save("images/presentation/"+name_tmp+"orig-sim.jpg", "JPEG")
        
        img_out = daltonize(daltonization_type, img_in, coldef_type)
        #img_out.show()
        img_out.save("images/presentation/"+name_tmp+"dalt.jpg", "JPEG")
        img_out_sim = simulate(simulation_type, img_out, coldef_type)
        #img_out_sim.show()
        img_out_sim.save("images/presentation/"+name_tmp+"dalt-sim.jpg", "JPEG")
    
def test4():
    best = [43,46,48,49,51,52,53,56]
    
    name = "example"
    coldef_type = "d"
    simulation_type = "vienot-adjusted"
    daltonization_type = "kotera"
    size = 128, 128
    
    for b in best:
        name_tmp = name+str(b)
        
        img_in = Image.open("images/"+name_tmp+".jpg")
        
        img_in.thumbnail(size, Image.ANTIALIAS)
        img_in.save("images/best/"+name_tmp+"orig.jpg", "JPEG")
        #img_in.show()
        img_in_sim = simulate(simulation_type, img_in, coldef_type)
        #img_in_sim.show()
        img_in_sim.save("images/best/"+name_tmp+"orig-sim.jpg", "JPEG")
        
        img_out = daltonize(daltonization_type, img_in, coldef_type)
        #img_out.show()
        img_out.save("images/best/"+name_tmp+"dalt.jpg", "JPEG")
        img_out_sim = simulate(simulation_type, img_out, coldef_type)
        #img_out_sim.show()
        img_out_sim.save("images/best/"+name_tmp+"dalt-sim.jpg", "JPEG")
                
        
def test3():
    
    name = "example10"
    simulation_type = "vienot-adjusted"
    coldef_type = "d"
    size = 512, 512
    
    input_tab, output_tab = makeSimulationLookupTable(simulation_type, coldef_type,4)
    #print input_tab, output_tab
    
    if True:
        img_in = Image.open("images/"+name+".jpg")
        img_in.thumbnail(size, Image.ANTIALIAS)
        #img_in.show()
        
        sRGB_in = colour.data.Data(colour.space.srgb,numpy.asarray(img_in)/255.)
        
        t = time.time()
        img_lut = lookup(img_in, input_tab, output_tab)
        print(time.time()-t)
        img_lut.show()
        
        sRGB_lut = colour.data.Data(colour.space.srgb,numpy.asarray(img_lut)/255.)
    
        t = time.time()
        img_sim = simulate(simulation_type,img_in,coldef_type)
        print(time.time()-t)
        img_sim.show()
        
        sRGB_sim = colour.data.Data(colour.space.srgb,numpy.asarray(img_sim)/255.)
        
        diff = colour.metric.dE_E(sRGB_in, sRGB_lut)
        print(diff)
        #print numpy.shape(diff)
        
        import pylab
        a = Image.fromarray(diff)
        #print numpy.max(diff)
        a.show()
        a = a.convert('RGB')
        a.save("./images/difference.jpg","JPEG")
        pylab.imshow(diff)
    else:
        import os
        size = 512,512
        directory = os.path.join("./images/test2/")
        for root,dirs,files in os.walk(directory):
            for file in files:
                if file.endswith('.jpg'):
                    dir = os.path.join(directory,file)
                    print(dir)
                    img_in = Image.open(dir)        
                    t = time.time()
                    img_lut = lookup(img_in, input_tab, output_tab)
                    #print time.time()-t
                    img_lut.thumbnail(size, Image.ANTIALIAS)
                    img_lut.show()
    


def test2():
    
    name = "example48"
    coldef_type = "d"
    simulation_type = "IPT"
    daltonization_type = "kotera"
    size = 512, 512
    
    img_in = Image.open("images/"+name+".jpg")
    img_in.thumbnail(size, Image.ANTIALIAS)
    img_in.show()
    img_in_sim = simulate(simulation_type, img_in, coldef_type)
    img_in_sim.show()
    
    img_out = daltonize(daltonization_type, img_in, coldef_type)
    img_out.show()
    img_out_sim = simulate(simulation_type, img_out, coldef_type)
    img_out_sim.show()


def test1():
    name = "test18"
    
    im = Image.open("images/"+name+".jpg")
    #im.show()
    simulation_type = "vienot"
    coldef_strength = 1.0
    
    for coldef_type in settings.coldef_types:
        im_sim = simulate(simulation_type,im,coldef_type,coldef_strength)
        #im_sim.show()
        im_sim.save(name+"-"+simulation_type+"-"+coldef_type+".jpg")
        print(coldef_type + " simulation done")
        
